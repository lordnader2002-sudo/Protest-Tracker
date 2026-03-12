# file: scripts/Simon OIC Intel - Protest Tracker Script v9.1.py
"""
Protest Tracker v9.1 (with Seen Store)
- General protests via Mobilize (ZIP-deduped querying)
- NoKings sheet (30-day default) from:
  - Mobilize org-filtered (No Kings)
  - Action Network scraping (seed pages -> event pages), with geocoding + radius matching

Seen store:
- Persists to seen_events.json (default next to script)
- Adds columns: Event Key, Is New, First Seen
- Highlights new rows in Excel (conditional formatting)
- Does NOT skip seen rows (exports them but marks Is New = FALSE)

Dependencies:
  pip install requests pandas openpyxl geopy
"""

from __future__ import annotations

import argparse
import datetime as dt
import html
import json
import os
import random
import re
import sys
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple, Union
from urllib.parse import urljoin, urlparse

import pandas as pd
import requests
from geopy.distance import geodesic
from geopy.geocoders import Nominatim


# -----------------------------
# Defaults
# -----------------------------
MOBILIZE_EVENTS_URL = "https://api.mobilize.us/v1/events"
MOBILIZE_ORGS_URL = "https://api.mobilize.us/v1/organizations"

DEFAULT_RADIUS_MILES = 3.0
DEFAULT_WINDOW_DAYS = 7
DEFAULT_NO_KINGS_WINDOW_DAYS = 30
DEFAULT_API_BUFFER_MILES = 8.0
DEFAULT_MIN_REQUEST_INTERVAL = 0.1

DEFAULT_SHEET_MAIN = "Protests"
DEFAULT_SHEET_MATCHES = "AllMatches"
DEFAULT_SHEET_NO_KINGS = "NoKings"

DEFAULT_OUTPUT_DIR = "."
DEFAULT_OUTPUT_PREFIX = "mobilize_protests"
DATE_STAMP_FMT_DAY = "%Y%d%m"  # YYYYDDMM
DATE_STAMP_FMT_MINUTE = "%Y%d%m_%H%M"  # YYYYDDMM_HHMM

DEFAULT_GEOCODE_CACHE = "geocode_cache.json"
DEFAULT_EVENT_GEOCODE_CACHE = "event_geocode_cache.json"
DEFAULT_SEEN_STORE = "seen_events.json"

DEFAULT_ACTION_NETWORK_SEEDS = [
    "https://actionnetwork.org/events/no-kings-3",
]
DEFAULT_ACTION_NETWORK_MAX_SEED_PAGES = 5
DEFAULT_ACTION_NETWORK_MAX_EVENT_LINKS = 500

DEFAULT_FILTER_CONFIG: Dict[str, Any] = {
    "include_event_types": ["RALLY", "VISIBILITY_EVENT", "SOLIDARITY_EVENT"],
    "exclude_event_types": [
        "PHONE_BANK",
        "TEXT_BANK",
        "MEETING",
        "TRAINING",
        "COMMUNITY",
        "FRIEND_TO_FRIEND_OUTREACH",
    ],
    "include_keywords": [
        "protest",
        "demonstration",
        "march",
        "rally",
        "walkout",
        "sit-in",
        "sit in",
        "picket",
        "strike",
        "boycott",
        "vigil",
    ],
    "exclude_keywords": [
        "phone bank",
        "phonebank",
        "text bank",
        "textbank",
        "training",
        "webinar",
        "call",
        "meeting",
    ],
    "match_mode": "any",
    "search_fields": ["title", "description"],
}

EXPORT_MAIN_COLUMNS = [
    "Source",
    "Organization",
    "Event Key",
    "Is New",
    "First Seen",
    "Last Seen",
    "Protest Name",
    "Date",
    "Time",
    "Location",
    "Nearest Property",
    "Nearest Property ID",
    "Distance to Nearest Property (miles)",
    "Event URL",
    "Event Type",
    "Event ID",
    "Timeslot Start (epoch)",
    "Matched Properties (within radius)",
]

EXPORT_MATCH_COLUMNS = [
    "Source",
    "Organization",
    "Event Key",
    "Is New",
    "First Seen",
    "Last Seen",
    "Protest Name",
    "Date",
    "Time",
    "Location",
    "Property",
    "Property ID",
    "Distance to Property (miles)",
    "Event URL",
    "Event Type",
    "Event ID",
    "Timeslot Start (epoch)",
]

# Columns dropped from Excel output (kept internally for processing/dedup)
_OUTPUT_HIDE_MAIN = {
    "Source", "Organization", "Nearest Property ID",
    "Event ID", "Timeslot Start (epoch)", "Matched Properties (within radius)",
}
_OUTPUT_HIDE_MATCH = {
    "Source", "Organization", "Property ID",
    "Event ID", "Timeslot Start (epoch)",
}

QueryParams = Union[Dict[str, str], Sequence[Tuple[str, str]]]


# -----------------------------
# Utilities
# -----------------------------
def script_dir() -> str:
    return os.path.dirname(os.path.abspath(__file__))


def default_properties_path() -> str:
    clean = os.path.join(script_dir(), "properties_clean.csv")
    raw = os.path.join(script_dir(), "properties.csv")
    return clean if os.path.exists(clean) else raw


def default_seen_store_path() -> str:
    return os.path.join(script_dir(), DEFAULT_SEEN_STORE)


def epoch_now() -> int:
    return int(time.time())


def compute_distance_miles(a: Tuple[float, float], b: Tuple[float, float]) -> float:
    return float(geodesic(a, b).miles)


def inject_stamp(path: str, stamp: str) -> str:
    directory, filename = os.path.split(path)
    if not filename:
        filename = "mobilize_protests.xlsx"
    base, ext = os.path.splitext(filename)
    ext = ext or ".xlsx"
    return os.path.join(directory or ".", f"{base}_{stamp}{ext}")


def build_output_path(args: argparse.Namespace) -> str:
    now = dt.datetime.now()
    stamp = now.strftime(DATE_STAMP_FMT_DAY if args.append else DATE_STAMP_FMT_MINUTE)
    if args.output:
        return inject_stamp(args.output, stamp)
    os.makedirs(args.output_dir, exist_ok=True)
    return os.path.join(args.output_dir, f"{args.output_prefix}_{stamp}.xlsx")


def safe_sheet_name(name: str, used: set[str]) -> str:
    cleaned = re.sub(r"[:\\/\\?\\*\\[\\]]", " ", name).strip()
    cleaned = re.sub(r"\s+", " ", cleaned)
    base = cleaned[:31] if cleaned else "Property"
    candidate = base
    i = 1
    while candidate in used or len(candidate) == 0:
        suffix = f"_{i}"
        candidate = (base[: 31 - len(suffix)] + suffix) if len(base) + len(suffix) > 31 else (base + suffix)
        i += 1
    used.add(candidate)
    return candidate


def autosize_worksheet_columns(writer: pd.ExcelWriter, sheet_name: str) -> None:
    ws = writer.sheets[sheet_name]
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 70)


def fmt_secs(s: float) -> str:
    s = int(round(max(0.0, s)))
    h, rem = divmod(s, 3600)
    m, sec = divmod(rem, 60)
    if h:
        return f"{h}:{m:02d}:{sec:02d}"
    return f"{m:02d}:{sec:02d}"


def render_progress(label: str, done: int, total: int, start_ts: float, failed: int, last_label: str) -> str:
    total = max(1, total)
    done = max(0, min(done, total))
    frac = done / total
    width = 24
    filled = int(frac * width)
    bar = "=" * filled + "-" * (width - filled)

    elapsed = max(0.0, time.time() - start_ts)
    rate = done / elapsed if elapsed > 0 else 0.0
    eta = (total - done) / rate if rate > 0 else 0.0
    pct = 100.0 * frac

    return (
        f"{label}: [{bar}] {pct:6.2f}% ({done}/{total}) "
        f"ETA {fmt_secs(eta)} | failed={failed} | last={last_label}"
    )


# -----------------------------
# Seen store
# -----------------------------
def load_seen_store(path: str) -> Dict[str, Dict[str, Any]]:
    try:
        if not os.path.exists(path):
            return {}
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def save_seen_store(path: str, store: Dict[str, Dict[str, Any]]) -> None:
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    tmp = f"{path}.tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(store, f, indent=2, sort_keys=True)
    os.replace(tmp, path)


def epoch_to_iso(ts: int) -> str:
    try:
        return dt.datetime.fromtimestamp(int(ts)).isoformat(timespec="seconds")
    except Exception:
        return ""


def build_event_key(source: str, event_id: str, event_url: str, timeslot_start: int) -> str:
    """
    Stable per-timeslot key across runs/sources.
    """
    source = (source or "").strip()
    event_id = (event_id or "").strip()
    event_url = (event_url or "").strip()
    ts = int(timeslot_start or 0)

    if source.lower().startswith("nokings-actionnetwork") or event_id.startswith("actionnetwork:"):
        base = event_id if event_id else f"actionnetwork:{event_url}"
        return f"{base}:{ts}"

    # Mobilize (General or NoKings-Mobilize)
    base = f"mobilize:{event_id}" if event_id else f"mobilize_url:{event_url}"
    return f"{base}:{ts}"


def apply_seen_flags(
    df: pd.DataFrame,
    store: Dict[str, Dict[str, Any]],
    now_ts: int,
) -> pd.DataFrame:
    """
    Adds/updates: Event Key, Is New, First Seen, Last Seen.
    Updates store in-place (first_seen/last_seen).
    """
    if df.empty:
        for col in ["Event Key", "Is New", "First Seen", "Last Seen"]:
            if col not in df.columns:
                df[col] = []
        return df

    event_keys: List[str] = []
    is_new_flags: List[bool] = []
    first_seen_vals: List[str] = []
    last_seen_vals: List[str] = []

    for _, row in df.iterrows():
        source = str(row.get("Source", "") or "")
        event_id = str(row.get("Event ID", "") or "")
        event_url = str(row.get("Event URL", "") or "")
        ts = int(row.get("Timeslot Start (epoch)", 0) or 0)

        key = build_event_key(source, event_id, event_url, ts)
        event_keys.append(key)

        rec = store.get(key)
        if not rec:
            store[key] = {"first_seen": now_ts, "last_seen": now_ts}
            is_new_flags.append(True)
            first_seen_vals.append(epoch_to_iso(now_ts))
            last_seen_vals.append(epoch_to_iso(now_ts))
        else:
            rec["last_seen"] = now_ts
            is_new_flags.append(False)
            first_seen_vals.append(epoch_to_iso(int(rec.get("first_seen") or now_ts)))
            last_seen_vals.append(epoch_to_iso(now_ts))

    df = df.copy()
    df["Event Key"] = event_keys
    df["Is New"] = is_new_flags
    df["First Seen"] = first_seen_vals
    df["Last Seen"] = last_seen_vals
    return df


def apply_highlight_new_rows(writer: pd.ExcelWriter, sheet_name: str) -> None:
    """
    Conditional format entire row when 'Is New' is TRUE.
    """
    try:
        from openpyxl.styles import PatternFill
        from openpyxl.formatting.rule import FormulaRule
        from openpyxl.utils import get_column_letter
    except Exception:
        return

    ws = writer.sheets.get(sheet_name)
    if ws is None or ws.max_row < 2 or ws.max_column < 1:
        return

    # find Is New column
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    try:
        is_new_idx = headers.index("Is New") + 1
    except ValueError:
        return

    is_new_col = get_column_letter(is_new_idx)
    start_row = 2
    end_row = ws.max_row
    start_col = 1
    end_col = ws.max_column

    fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # light yellow
    # Formula evaluated per row; lock column, not row.
    formula = f"${is_new_col}{start_row}=TRUE"
    rule = FormulaRule(formula=[formula], fill=fill)

    rng = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
    ws.conditional_formatting.add(rng, rule)


def apply_sheet_formatting(writer: pd.ExcelWriter, sheet_name: str) -> None:
    """Freeze top row, bold key columns, and shade rows by distance to property."""
    try:
        from openpyxl.styles import PatternFill, Font
    except Exception:
        return

    ws = writer.sheets.get(sheet_name)
    if ws is None or ws.max_row < 2:
        return

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    # Freeze top row
    ws.freeze_panes = "A2"

    # Bold specific columns (header + all data rows)
    bold_names = {
        "Protest Name", "Date", "Time",
        "Nearest Property", "Property",
        "Distance to Nearest Property (miles)", "Distance to Property (miles)",
    }
    bold_font = Font(bold=True)
    for col_idx, header in enumerate(headers, start=1):
        if header in bold_names:
            for row_idx in range(1, ws.max_row + 1):
                ws.cell(row=row_idx, column=col_idx).font = bold_font

    # Distance-based row shading
    dist_col_idx = None
    for name in ("Distance to Nearest Property (miles)", "Distance to Property (miles)"):
        if name in headers:
            dist_col_idx = headers.index(name) + 1
            break

    if dist_col_idx is None:
        return

    fill_red   = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # < 1 mile
    fill_amber = PatternFill(start_color="FFE5B4", end_color="FFE5B4", fill_type="solid")  # 1-2 miles
    fill_green = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")  # > 2 miles

    for row_idx in range(2, ws.max_row + 1):
        try:
            dist = float(ws.cell(row=row_idx, column=dist_col_idx).value)
        except (TypeError, ValueError):
            continue
        fill = fill_red if dist < 1.0 else (fill_amber if dist < 2.0 else fill_green)
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill

    # Italic gray text for duplicate rows (applied after bold so bold is preserved)
    dup_col_idx = headers.index("Is Duplicate") + 1 if "Is Duplicate" in headers else None
    if dup_col_idx:
        for row_idx in range(2, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=dup_col_idx).value
            if val is True or str(val).upper() == "TRUE":
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.font = Font(bold=cell.font.bold, italic=True, color="999999")


def mark_duplicates(df: pd.DataFrame) -> pd.DataFrame:
    """Add 'Is Duplicate' column: True for rows sharing the same title+date+location."""
    df = df.copy()
    title_col = "Protest Name" if "Protest Name" in df.columns else None
    date_col = "Date" if "Date" in df.columns else None
    loc_col = "Location" if "Location" in df.columns else None

    if title_col and date_col and loc_col:
        def _norm(s: object) -> str:
            return " ".join(str(s or "").lower().split())
        fp = (df[title_col].map(_norm) + "|" +
              df[date_col].map(_norm) + "|" +
              df[loc_col].map(_norm))
        df["Is Duplicate"] = fp.duplicated(keep="first")
    else:
        df["Is Duplicate"] = False

    # Place "Is Duplicate" immediately after "Is New"
    if "Is New" in df.columns:
        cols = [c for c in df.columns if c != "Is Duplicate"]
        cols.insert(cols.index("Is New") + 1, "Is Duplicate")
        df = df[cols]

    return df


# -----------------------------
# Rate limiter
# -----------------------------
class GlobalRateLimiter:
    def __init__(self, min_interval_s: float) -> None:
        self._min = float(min_interval_s)
        self._lock = threading.Lock()
        self._next_ok = 0.0

    def wait(self) -> None:
        if self._min <= 0:
            return
        with self._lock:
            now = time.time()
            if now < self._next_ok:
                time.sleep(self._next_ok - now)
            self._next_ok = time.time() + self._min + random.uniform(0.0, 0.15)


# -----------------------------
# CLI
# -----------------------------
def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Protest Tracker v9.1 (General + NoKings + Seen Store)")
    p.add_argument("--properties", required=False, default=default_properties_path())
    p.add_argument("--radius-miles", type=float, default=DEFAULT_RADIUS_MILES)
    p.add_argument("--window-days", type=int, default=DEFAULT_WINDOW_DAYS)
    p.add_argument("--api-buffer-miles", type=float, default=DEFAULT_API_BUFFER_MILES)

    p.add_argument("--min-request-interval", type=float, default=DEFAULT_MIN_REQUEST_INTERVAL)
    p.add_argument("--workers", type=int, default=8)

    p.add_argument("--output", default=None)
    p.add_argument("--output-dir", default=DEFAULT_OUTPUT_DIR)
    p.add_argument("--output-prefix", default=DEFAULT_OUTPUT_PREFIX)

    p.add_argument("--append", action="store_true")
    p.add_argument("--keep-all-matches", action="store_true")
    p.add_argument("--per-property-output", choices=["none", "sheets"], default="none")

    p.add_argument("--filter-config", default=None)
    p.add_argument("--include-non-protest", action="store_true")

    p.add_argument("--geocode-missing", action="store_true")
    p.add_argument("--geocode-cache", default=DEFAULT_GEOCODE_CACHE)
    p.add_argument("--event-geocode-cache", default=DEFAULT_EVENT_GEOCODE_CACHE)
    p.add_argument("--geocode-rate-seconds", type=float, default=1.2)

    p.add_argument("--per-page", type=int, default=200)
    p.add_argument("--timeout", type=int, default=60)

    p.add_argument("--progress", action="store_true")
    p.add_argument("--no-autosize", action="store_true", help="Skip Excel column autosizing (faster).")

    p.add_argument("--validate-properties", nargs="?", const="bad_rows.csv", default=None)
    p.add_argument("--rewrite-properties", default=None)

    # Seen store
    p.add_argument(
        "--seen-store",
        default=default_seen_store_path(),
        help="JSON file to store first_seen/last_seen for events (default: seen_events.json next to script).",
    )
    p.add_argument(
        "--no-highlight-new",
        action="store_true",
        help="Disable Excel highlighting for new events.",
    )

    # No Kings
    p.add_argument("--no-kings", action="store_true")
    p.add_argument("--no-kings-window-days", type=int, default=DEFAULT_NO_KINGS_WINDOW_DAYS)
    p.add_argument("--no-kings-slug", default="nokings")
    p.add_argument("--no-kings-org-id", type=int, default=None)

    # Action Network (No Kings)
    p.add_argument(
        "--action-network-seeds",
        default=",".join(DEFAULT_ACTION_NETWORK_SEEDS),
        help="Comma-separated Action Network seed URLs to scrape for No Kings events.",
    )
    p.add_argument("--action-network-max-seed-pages", type=int, default=DEFAULT_ACTION_NETWORK_MAX_SEED_PAGES)
    p.add_argument("--action-network-max-event-links", type=int, default=DEFAULT_ACTION_NETWORK_MAX_EVENT_LINKS)

    return p.parse_args()


# -----------------------------
# Properties ingestion / cleaning
# -----------------------------
def read_properties(path: str) -> pd.DataFrame:
    ext = os.path.splitext(path.lower())[1]
    if ext == ".csv":
        df = pd.read_csv(path, dtype=str).fillna("")
    elif ext in {".xlsx", ".xls"}:
        df = pd.read_excel(path, dtype=str, engine="openpyxl").fillna("")
    else:
        raise ValueError("properties file must be .csv or .xlsx/.xls")

    df.columns = [str(c).strip().lower().replace(" ", "_") for c in df.columns]
    df = df.rename(
        columns={
            "zipcode": "postal_code",
            "zip": "postal_code",
            "postal": "postal_code",
            "postalcode": "postal_code",
            "queryzip": "query_zipcode",
            "queryzipcode": "query_zipcode",
            "query_zip": "query_zipcode",
        }
    )
    return df


def normalize_us_zip(zipcode: str) -> str:
    digits = re.sub(r"\D", "", zipcode or "")
    if not digits:
        return ""
    if len(digits) <= 5:
        return digits.zfill(5)
    if len(digits) == 9:
        return digits[:5]
    return digits


def extract_us_zip_from_text(text: str) -> str:
    matches = re.findall(r"\b(\d{5})(?:-\d{4})?\b", text or "")
    return matches[-1] if matches else ""


def infer_postal_and_query_zip(postal_raw: str, query_raw: str, address: str) -> Tuple[str, str, str]:
    postal_raw = (postal_raw or "").strip()
    query_raw = (query_raw or "").strip()
    address = (address or "").strip()

    provided_us = normalize_us_zip(postal_raw)
    postal_code = provided_us if re.fullmatch(r"\d{5}", provided_us) else postal_raw

    query_zip = normalize_us_zip(query_raw)
    if re.fullmatch(r"\d{5}", query_zip):
        return postal_code, query_zip, "query_zipcode from explicit column"

    if re.fullmatch(r"\d{5}", provided_us):
        return postal_code, provided_us, "query_zipcode from provided postal_code"

    inferred = normalize_us_zip(extract_us_zip_from_text(address))
    if re.fullmatch(r"\d{5}", inferred):
        return postal_code, inferred, "query_zipcode inferred from address"

    return postal_code, "", "missing query_zipcode"


def validate_properties_to_csv(df: pd.DataFrame, out_path: str) -> int:
    required_cols = {"name"}
    missing_required = sorted([c for c in required_cols if c not in df.columns])
    if missing_required:
        pd.DataFrame(
            [{"row": 0, "property_id": "", "name": "", "issue": f"missing required column(s): {', '.join(missing_required)}"}]
        ).to_csv(out_path, index=False)
        return 1

    for col in ["property_id", "address", "postal_code", "query_zipcode", "lat", "lon"]:
        if col not in df.columns:
            df[col] = ""

    bad_rows: List[Dict[str, Any]] = []
    for i, row in df.fillna("").iterrows():
        row_num = int(i) + 2
        name = str(row.get("name", "")).strip()
        property_id = str(row.get("property_id", "")).strip() or f"PROP_{row_num:04d}"
        address = str(row.get("address", "")).strip()

        postal_raw = str(row.get("postal_code", "")).strip()
        query_raw = str(row.get("query_zipcode", "")).strip()
        postal_code, query_zip, _ = infer_postal_and_query_zip(postal_raw, query_raw, address)

        lat_s = str(row.get("lat", "")).strip()
        lon_s = str(row.get("lon", "")).strip()

        issues: List[str] = []
        if not name:
            issues.append("missing name")

        if not postal_code:
            issues.append("missing postal_code")
        else:
            if not re.fullmatch(r"\d{5}", normalize_us_zip(postal_raw)):
                issues.append("invalid postal_code (expected US 5-digit ZIP)")

        if not query_zip:
            issues.append("missing query_zipcode for Mobilize querying")
        elif not re.fullmatch(r"\d{5}", query_zip):
            issues.append(f"invalid query_zipcode: {query_zip!r}")

        has_lat = bool(lat_s)
        has_lon = bool(lon_s)
        if has_lat != has_lon:
            issues.append("lat/lon must be both filled or both blank")
        if has_lat and has_lon:
            try:
                float(lat_s)
                float(lon_s)
            except ValueError:
                issues.append(f"invalid lat/lon: {lat_s!r}, {lon_s!r}")

        if issues:
            bad_rows.append(
                {
                    "row": row_num,
                    "property_id": property_id,
                    "name": name,
                    "postal_code": postal_code,
                    "query_zipcode": query_zip,
                    "address": address,
                    "lat": lat_s,
                    "lon": lon_s,
                    "issue": "; ".join(issues),
                }
            )

    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    pd.DataFrame(bad_rows).to_csv(out_path, index=False)
    return len(bad_rows)


def rewrite_properties(df: pd.DataFrame, out_csv: str) -> None:
    for col in ["property_id", "name", "address", "postal_code", "query_zipcode", "lat", "lon"]:
        if col not in df.columns:
            df[col] = ""

    out = df.copy()
    norm_postals: List[str] = []
    norm_queries: List[str] = []
    notes_col: List[str] = []

    for _, row in out.fillna("").iterrows():
        postal_raw = str(row.get("postal_code", "")).strip()
        query_raw = str(row.get("query_zipcode", "")).strip()
        address = str(row.get("address", "")).strip()
        postal_code, query_zip, note = infer_postal_and_query_zip(postal_raw, query_raw, address)
        norm_postals.append(postal_code)
        norm_queries.append(query_zip)
        notes_col.append(note)

    out["postal_code"] = norm_postals
    out["query_zipcode"] = norm_queries
    out["rewrite_note"] = notes_col

    os.makedirs(os.path.dirname(out_csv) or ".", exist_ok=True)
    out.to_csv(out_csv, index=False)
    print(f"Wrote cleaned properties CSV: {out_csv}")


# -----------------------------
# JSON cache helpers
# -----------------------------
def load_json_cache(path: str) -> Dict[str, Dict[str, Any]]:
    if not os.path.exists(path):
        return {}
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def save_json_cache(path: str, cache: Dict[str, Dict[str, Any]]) -> None:
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    tmp = f"{path}.tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(cache, f, indent=2, sort_keys=True)
    os.replace(tmp, path)


def geocode_cached(
    *,
    cache: Dict[str, Dict[str, Any]],
    cache_key: str,
    address: str,
    geolocator: Nominatim,
    rate_seconds: float,
) -> Optional[Tuple[float, float]]:
    if cache_key in cache and "lat" in cache[cache_key] and "lon" in cache[cache_key]:
        return float(cache[cache_key]["lat"]), float(cache[cache_key]["lon"])

    loc = geolocator.geocode(address)
    if not loc:
        return None
    lat, lon = float(loc.latitude), float(loc.longitude)
    cache[cache_key] = {"lat": lat, "lon": lon, "ts": int(time.time())}
    time.sleep(rate_seconds)
    return lat, lon


# -----------------------------
# Property model
# -----------------------------
@dataclass(frozen=True)
class Property:
    property_id: str
    name: str
    address: str
    postal_code: str
    query_zipcode: str
    lat: float
    lon: float


def normalize_properties(
    df: pd.DataFrame,
    geocode_missing: bool,
    geocode_cache_path: str,
    geocode_rate_seconds: float,
) -> List[Property]:
    for col in ["property_id", "name", "address", "postal_code", "query_zipcode", "lat", "lon"]:
        if col not in df.columns:
            df[col] = ""

    df = df.fillna("")
    cache = load_json_cache(geocode_cache_path)
    geolocator = Nominatim(user_agent="company-protest-monitor/9.1 (contact: you@company.com)", timeout=20)

    props: List[Property] = []
    warnings: List[str] = []

    for idx, row in df.iterrows():
        name = str(row["name"]).strip()
        if not name:
            continue

        property_id = str(row.get("property_id", "")).strip() or f"PROP_{idx+1:04d}"
        address = str(row.get("address", "")).strip()

        postal_raw = str(row.get("postal_code", "")).strip()
        query_raw = str(row.get("query_zipcode", "")).strip()
        postal_code, query_zip, _ = infer_postal_and_query_zip(postal_raw, query_raw, address)

        lat_s = str(row.get("lat", "")).strip()
        lon_s = str(row.get("lon", "")).strip()

        lat: Optional[float] = None
        lon: Optional[float] = None

        if lat_s and lon_s:
            try:
                lat, lon = float(lat_s), float(lon_s)
            except ValueError:
                lat, lon = None, None

        if (lat is None or lon is None) and geocode_missing:
            if not address:
                warnings.append(f"{property_id} ({name}) missing address/lat/lon")
            else:
                coords = geocode_cached(
                    cache=cache,
                    cache_key=address,
                    address=address,
                    geolocator=geolocator,
                    rate_seconds=geocode_rate_seconds,
                )
                if coords:
                    lat, lon = coords
                else:
                    warnings.append(f"{property_id} ({name}) geocode failed: {address}")

        if lat is None or lon is None:
            warnings.append(f"{property_id} ({name}) missing lat/lon (skipped)")
            continue

        if not query_zip or not re.fullmatch(r"\d{5}", query_zip):
            warnings.append(f"{property_id} ({name}) missing/invalid query_zipcode (skipped)")
            continue

        if not postal_code:
            postal_code = query_zip

        props.append(
            Property(
                property_id=property_id,
                name=name,
                address=address,
                postal_code=postal_code,
                query_zipcode=query_zip,
                lat=lat,
                lon=lon,
            )
        )

    if geocode_missing:
        save_json_cache(geocode_cache_path, cache)

    if warnings:
        print("Warnings / skipped properties:", file=sys.stderr)
        for line in warnings[:120]:
            print(f" - {line}", file=sys.stderr)

    if not props:
        raise RuntimeError("No valid properties loaded. Fix properties file or enable --geocode-missing.")

    return props


def group_properties_by_query_zip(props: List[Property]) -> Dict[str, List[Property]]:
    grouped: Dict[str, List[Property]] = {}
    for p in props:
        grouped.setdefault(p.query_zipcode, []).append(p)
    return grouped


# -----------------------------
# Filtering
# -----------------------------
def load_filter_config(path: Optional[str]) -> Dict[str, Any]:
    if not path:
        cfg = DEFAULT_FILTER_CONFIG.copy()
    else:
        with open(path, "r", encoding="utf-8") as f:
            user_cfg = json.load(f)
        cfg = DEFAULT_FILTER_CONFIG.copy()
        cfg.update(user_cfg)

    cfg["include_event_types"] = [str(x).upper() for x in cfg.get("include_event_types", [])]
    cfg["exclude_event_types"] = [str(x).upper() for x in cfg.get("exclude_event_types", [])]
    cfg["include_keywords"] = [str(x).lower() for x in cfg.get("include_keywords", [])]
    cfg["exclude_keywords"] = [str(x).lower() for x in cfg.get("exclude_keywords", [])]
    cfg["match_mode"] = str(cfg.get("match_mode", "any")).lower()
    cfg["search_fields"] = [str(x).lower() for x in cfg.get("search_fields", ["title", "description"])]

    if cfg["match_mode"] not in {"any", "all"}:
        raise ValueError("filter config: match_mode must be 'any' or 'all'")
    if not set(cfg["search_fields"]).issubset({"title", "description"}):
        raise ValueError("filter config: search_fields must be subset of ['title','description']")
    return cfg


def strip_html_to_text(raw_html: str) -> str:
    if not raw_html:
        return ""
    raw_html = html.unescape(raw_html)
    raw_html = re.sub(r"<(script|style)[^>]*>.*?</\1>", " ", raw_html, flags=re.I | re.S)
    text = re.sub(r"<[^>]+>", " ", raw_html)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def build_search_blob(event: Dict[str, Any], fields: Sequence[str]) -> str:
    parts: List[str] = []
    if "title" in fields:
        parts.append(str(event.get("title") or ""))
    if "description" in fields:
        parts.append(strip_html_to_text(str(event.get("description") or "")))
    return " ".join(parts).lower()


def looks_like_protest(event: Dict[str, Any], cfg: Dict[str, Any]) -> bool:
    event_type = str(event.get("event_type") or "").strip().upper()
    if event_type and event_type in set(cfg.get("exclude_event_types", [])):
        return False

    blob = build_search_blob(event, cfg.get("search_fields", ["title", "description"]))
    if any(k in blob for k in cfg.get("exclude_keywords", [])):
        return False

    if event_type and event_type in set(cfg.get("include_event_types", [])):
        return True

    include_keywords = cfg.get("include_keywords", [])
    if not include_keywords:
        return False

    if cfg.get("match_mode", "any") == "all":
        return all(k in blob for k in include_keywords)
    return any(k in blob for k in include_keywords)


# -----------------------------
# Mobilize API
# -----------------------------
def request_with_retries(
    session: requests.Session,
    url: str,
    params: Any,
    timeout: int,
    limiter: GlobalRateLimiter,
) -> Dict[str, Any]:
    backoff = 1.0
    last_status: Optional[int] = None
    last_text: Optional[str] = None
    last_exc: Optional[Exception] = None

    for _ in range(10):
        try:
            limiter.wait()
            r = session.get(url, params=params, timeout=timeout)
            last_status = r.status_code
            last_text = (r.text or "")[:300]

            if r.status_code in {429, 500, 502, 503, 504}:
                ra = r.headers.get("Retry-After")
                if ra and ra.isdigit():
                    time.sleep(min(int(ra), 60))
                else:
                    time.sleep(backoff + random.uniform(0.0, 0.25))
                    backoff = min(backoff * 2, 30)
                continue

            r.raise_for_status()
            return r.json()

        except Exception as e:
            last_exc = e
            time.sleep(backoff + random.uniform(0.0, 0.25))
            backoff = min(backoff * 2, 30)

    raise RuntimeError(
        f"Request failed after retries: status={last_status}, exc={last_exc!r}, body={last_text!r}"
    )


def fetch_all_events(
    session: requests.Session,
    params: QueryParams,
    timeout: int,
    limiter: GlobalRateLimiter,
) -> Iterable[Dict[str, Any]]:
    url: Optional[str] = MOBILIZE_EVENTS_URL
    while url:
        payload = request_with_retries(session, url, params if url == MOBILIZE_EVENTS_URL else None, timeout, limiter)
        for item in payload.get("data") or []:
            yield item
        url = payload.get("next")


def build_location_string_mobilize(event_location: Optional[Dict[str, Any]]) -> str:
    if not event_location:
        return ""
    venue = (event_location.get("venue") or "").strip()
    lines = event_location.get("address_lines") or []
    lines = [str(x).strip() for x in lines if str(x).strip()]
    locality = (event_location.get("locality") or "").strip()
    region = (event_location.get("region") or "").strip()
    postal = (event_location.get("postal_code") or "").strip()

    parts: List[str] = []
    if venue and "private" not in venue.lower():
        parts.append(venue)
    if lines:
        parts.append(", ".join(lines))
    city_state = ", ".join([p for p in [locality, region] if p]).strip()
    if postal:
        city_state = f"{city_state} {postal}".strip()
    if city_state:
        parts.append(city_state)
    return " | ".join(parts) if parts else ""


def pick_earliest_timeslot_in_window(
    timeslots: List[Dict[str, Any]], start_epoch: int, end_epoch: int
) -> Optional[Dict[str, Any]]:
    candidates = [
        t
        for t in timeslots
        if isinstance(t.get("start_date"), int)
        and start_epoch <= int(t["start_date"]) < end_epoch
    ]
    if not candidates:
        return None
    return sorted(candidates, key=lambda t: int(t["start_date"]))[0]


def format_timeslot_local(start_epoch: int, tz_name: str) -> Tuple[str, str]:
    try:
        from zoneinfo import ZoneInfo

        tz = ZoneInfo(tz_name)
        local_dt = dt.datetime.fromtimestamp(start_epoch, tz=dt.timezone.utc).astimezone(tz)
    except Exception:
        local_dt = dt.datetime.utcfromtimestamp(start_epoch)
    return local_dt.strftime("%Y-%m-%d"), local_dt.strftime("%I:%M %p").lstrip("0")


def resolve_mobilize_org_by_slug(slug: str, timeout: int, limiter: GlobalRateLimiter) -> Tuple[int, str]:
    session = requests.Session()
    url: Optional[str] = MOBILIZE_ORGS_URL
    params: Optional[List[Tuple[str, str]]] = [("per_page", "200")]

    for _ in range(200):
        payload = request_with_retries(session, url or MOBILIZE_ORGS_URL, params, timeout, limiter)
        data = payload.get("data") or []
        for org in data:
            if str(org.get("slug") or "").strip().lower() == slug.strip().lower():
                return int(org["id"]), str(org.get("name") or "").strip()
        nxt = payload.get("next")
        if not nxt:
            break
        url = nxt
        params = None

    raise RuntimeError(f"Could not resolve Mobilize organization for slug={slug!r}.")


def query_mobilize_events_for_zip(
    query_zipcode: str,
    now_epoch: int,
    end_epoch: int,
    api_max_dist: float,
    per_page: int,
    timeout: int,
    include_non_protest: bool,
    filter_cfg: Dict[str, Any],
    limiter: GlobalRateLimiter,
    organization_id: Optional[int],
) -> List[Dict[str, Any]]:
    session = requests.Session()
    params: List[Tuple[str, str]] = [
        ("zipcode", query_zipcode),
        ("max_dist", str(api_max_dist)),
        ("is_virtual", "false"),
        ("per_page", str(per_page)),
        ("timeslot_start", f"gte_{now_epoch}"),
        ("timeslot_start", f"lt_{end_epoch}"),
    ]
    if organization_id is not None:
        params.append(("organization_id", str(organization_id)))

    out: List[Dict[str, Any]] = []
    for event in fetch_all_events(session, params, timeout=timeout, limiter=limiter):
        if not include_non_protest and not looks_like_protest(event, filter_cfg):
            continue

        location = event.get("location") or {}
        loc_geo = (location.get("location") or {}) if isinstance(location, dict) else {}
        lat = loc_geo.get("latitude")
        lon = loc_geo.get("longitude")
        if lat is None or lon is None:
            continue

        timeslot = pick_earliest_timeslot_in_window(event.get("timeslots") or [], now_epoch, end_epoch)
        if not timeslot:
            continue

        start_epoch = int(timeslot["start_date"])
        tz_name = (event.get("timezone") or "UTC").strip()
        date_str, time_str = format_timeslot_local(start_epoch, tz_name)

        out.append(
            {
                "event_id": str(event.get("id") or ""),
                "event_type": str(event.get("event_type") or ""),
                "title": str(event.get("title") or ""),
                "browser_url": str(event.get("browser_url") or ""),
                "location_str": build_location_string_mobilize(location),
                "event_lat": float(lat),
                "event_lon": float(lon),
                "timeslot_start": start_epoch,
                "date": date_str,
                "time": time_str,
            }
        )
    return out


# -----------------------------
# Action Network scraping (public pages)
# -----------------------------
def parse_iso_to_epoch(s: str) -> Optional[int]:
    if not s:
        return None
    s = s.strip()
    try:
        d = dt.datetime.fromisoformat(s.replace("Z", "+00:00"))
        if d.tzinfo is None:
            return int(d.replace(tzinfo=dt.timezone.utc).timestamp())
        return int(d.timestamp())
    except Exception:
        return None


def _extract_json_ld_events(html_text: str) -> List[Dict[str, Any]]:
    blocks = re.findall(
        r'<script[^>]+type=["\']application/ld\+json["\'][^>]*>(.*?)</script>',
        html_text,
        flags=re.I | re.S,
    )
    events: List[Dict[str, Any]] = []
    for b in blocks:
        b = b.strip()
        if not b:
            continue
        try:
            data = json.loads(b)
        except Exception:
            continue

        candidates: List[Any]
        if isinstance(data, list):
            candidates = data
        else:
            candidates = [data]

        for item in candidates:
            if not isinstance(item, dict):
                continue
            t = item.get("@type")
            if isinstance(t, list):
                is_event = any(str(x).lower() == "event" for x in t)
            else:
                is_event = str(t).lower() == "event"
            if is_event:
                events.append(item)
    return events


def _format_location_from_schema(location_obj: Any) -> str:
    if not isinstance(location_obj, dict):
        return ""
    place_name = str(location_obj.get("name") or "").strip()
    addr = location_obj.get("address")
    if isinstance(addr, dict):
        street = str(addr.get("streetAddress") or "").strip()
        city = str(addr.get("addressLocality") or "").strip()
        region = str(addr.get("addressRegion") or "").strip()
        postal = str(addr.get("postalCode") or "").strip()
        parts = [p for p in [street, ", ".join([x for x in [city, region] if x]).strip(), postal] if p]
        addr_str = " ".join(parts).strip()
    else:
        addr_str = str(addr or "").strip()

    out_parts = []
    if place_name:
        out_parts.append(place_name)
    if addr_str:
        out_parts.append(addr_str)
    return " | ".join(out_parts)


def _extract_address_for_geocode(location_obj: Any) -> str:
    if not isinstance(location_obj, dict):
        return ""
    addr = location_obj.get("address")
    if isinstance(addr, dict):
        street = str(addr.get("streetAddress") or "").strip()
        city = str(addr.get("addressLocality") or "").strip()
        region = str(addr.get("addressRegion") or "").strip()
        postal = str(addr.get("postalCode") or "").strip()
        country = str(addr.get("addressCountry") or "").strip()
        parts = [street, city, region, postal, country]
        return ", ".join([p for p in parts if p]).strip(", ")
    return str(addr or "").strip()


def scrape_action_network_event_page(url: str, timeout: int) -> Optional[Dict[str, Any]]:
    r = requests.get(url, timeout=timeout, headers={"User-Agent": "Mozilla/5.0"})
    r.raise_for_status()
    html_text = r.text or ""

    ld_events = _extract_json_ld_events(html_text)
    if not ld_events:
        return None

    ev = ld_events[0]
    title = str(ev.get("name") or ev.get("headline") or "").strip()
    desc = str(ev.get("description") or "").strip()
    start_s = str(ev.get("startDate") or "").strip()
    end_s = str(ev.get("endDate") or "").strip()

    start_epoch = parse_iso_to_epoch(start_s)
    if not start_epoch:
        return None

    location_obj = ev.get("location") or {}
    location_str = _format_location_from_schema(location_obj)
    addr_for_geo = _extract_address_for_geocode(location_obj)

    local_dt = dt.datetime.fromtimestamp(start_epoch)
    return {
        "event_id": f"actionnetwork:{url}",
        "event_type": "ACTION_NETWORK_EVENT",
        "title": title or "No Kings Event",
        "browser_url": url,
        "location_str": location_str,
        "address_for_geocode": addr_for_geo,
        "timeslot_start": start_epoch,
        "date": local_dt.strftime("%Y-%m-%d"),
        "time": local_dt.strftime("%I:%M %p").lstrip("0"),
        "description": desc,
        "end_epoch": parse_iso_to_epoch(end_s) if end_s else None,
    }


def scrape_action_network_seed_for_event_links(
    seed_url: str,
    timeout: int,
    max_pages: int,
    max_links: int,
) -> List[str]:
    seen: set[str] = set()
    links: List[str] = []

    current = seed_url
    for _ in range(max_pages):
        r = requests.get(current, timeout=timeout, headers={"User-Agent": "Mozilla/5.0"})
        r.raise_for_status()
        html_text = r.text or ""

        hrefs = re.findall(r'href=["\']([^"\']+)["\']', html_text, flags=re.I)
        for h in hrefs:
            abs_url = urljoin(current, h).split("#")[0]
            if "actionnetwork.org/events/" not in abs_url:
                continue
            if abs_url.rstrip("/") == seed_url.rstrip("/"):
                continue
            if abs_url in seen:
                continue
            seen.add(abs_url)
            links.append(abs_url)
            if len(links) >= max_links:
                return links

        next_url = None
        m = re.search(r'href=["\']([^"\']+)["\'][^>]*>\s*Next\s*<', html_text, flags=re.I)
        if m:
            next_url = urljoin(current, m.group(1))
        else:
            parsed = urlparse(current)
            qs = parsed.query or ""
            pm = re.search(r"(?:^|&)page=(\d+)(?:&|$)", qs)
            if pm:
                page_n = int(pm.group(1)) + 1
                next_url = re.sub(r"page=\d+", f"page={page_n}", current)
            else:
                joiner = "&" if ("?" in current) else "?"
                next_url = f"{current}{joiner}page=2"

        if not next_url or next_url == current:
            break
        current = next_url

    return links


def collect_action_network_events(
    seeds: List[str],
    now_epoch: int,
    end_epoch: int,
    timeout: int,
    max_seed_pages: int,
    max_event_links: int,
    show_progress: bool,
) -> List[Dict[str, Any]]:
    event_urls: List[str] = []
    for seed in seeds:
        try:
            event_urls.extend(
                scrape_action_network_seed_for_event_links(
                    seed_url=seed,
                    timeout=timeout,
                    max_pages=max_seed_pages,
                    max_links=max_event_links,
                )
            )
        except Exception as e:
            print(f"[ActionNetwork] seed failed: {seed}: {e}", file=sys.stderr)

    uniq_urls = list(dict.fromkeys(event_urls))
    if not uniq_urls:
        return []

    start_ts = time.time()
    done = 0
    failed = 0
    events: List[Dict[str, Any]] = []
    lock = threading.Lock()

    def _fetch_one(url: str) -> Optional[Dict[str, Any]]:
        return scrape_action_network_event_page(url, timeout=timeout)

    with ThreadPoolExecutor(max_workers=10) as ex:
        futures = {ex.submit(_fetch_one, url): url for url in uniq_urls}
        for fut in as_completed(futures):
            url = futures[fut]
            try:
                ev = fut.result()
                if ev and (now_epoch <= int(ev["timeslot_start"]) < end_epoch):
                    with lock:
                        events.append(ev)
            except Exception as e:
                with lock:
                    failed += 1
                print(f"[ActionNetwork] event page failed: {url}: {e}", file=sys.stderr)
            finally:
                with lock:
                    done += 1
                if show_progress:
                    line = render_progress("AN scrape", done, len(uniq_urls), start_ts, failed, url[-28:])
                    sys.stdout.write("\r" + line)
                    sys.stdout.flush()

    if show_progress:
        sys.stdout.write("\n")
        sys.stdout.flush()

    return events


# -----------------------------
# Matching + export shaping
# -----------------------------
def build_matches_for_events(
    events: List[Dict[str, Any]],
    props: List[Property],
    radius_miles: float,
    source_label: str,
    org_name: str,
    event_geocode_cache_path: str,
    geocode_rate_seconds: float,
    show_progress: bool,
) -> pd.DataFrame:
    geolocator = Nominatim(user_agent="company-protest-monitor/9.1-events (contact: you@company.com)", timeout=20)
    ev_cache = load_json_cache(event_geocode_cache_path)

    start_ts = time.time()
    done = 0
    failed = 0
    rows: List[Dict[str, Any]] = []

    for ev in events:
        try:
            ev_lat = ev.get("event_lat")
            ev_lon = ev.get("event_lon")

            if ev_lat is None or ev_lon is None:
                addr = str(ev.get("address_for_geocode") or "").strip()
                if addr:
                    coords = geocode_cached(
                        cache=ev_cache,
                        cache_key=addr,
                        address=addr,
                        geolocator=geolocator,
                        rate_seconds=geocode_rate_seconds,
                    )
                    if coords:
                        ev_lat, ev_lon = coords
                        ev["event_lat"] = float(ev_lat)
                        ev["event_lon"] = float(ev_lon)

            if ev_lat is None or ev_lon is None:
                failed += 1
                continue

            ep = (float(ev_lat), float(ev_lon))
            for prop in props:
                pp = (prop.lat, prop.lon)
                dist = compute_distance_miles(pp, ep)
                if dist > radius_miles:
                    continue
                rows.append(
                    {
                        "Source": source_label,
                        "Organization": org_name,
                        "Protest Name": ev.get("title", ""),
                        "Date": ev.get("date", ""),
                        "Time": ev.get("time", ""),
                        "Location": ev.get("location_str", ""),
                        "Property": prop.name,
                        "Property ID": prop.property_id,
                        "Distance to Property (miles)": round(dist, 2),
                        "Event URL": ev.get("browser_url", ""),
                        "Event Type": ev.get("event_type", ""),
                        "Event ID": ev.get("event_id", ""),
                        "Timeslot Start (epoch)": int(ev.get("timeslot_start") or 0),
                    }
                )
        finally:
            done += 1
            if show_progress:
                last = str(ev.get("browser_url") or "")[-28:] or str(ev.get("title") or "")[:28]
                line = render_progress("AN match/geocode", done, len(events), start_ts, failed, last)
                sys.stdout.write("\r" + line)
                sys.stdout.flush()

    if show_progress:
        sys.stdout.write("\n")
        sys.stdout.flush()

    save_json_cache(event_geocode_cache_path, ev_cache)

    if not rows:
        return pd.DataFrame(columns=EXPORT_MATCH_COLUMNS)

    df = pd.DataFrame(rows)
    # placeholders; seen will be applied later
    df["Event Key"] = ""
    df["Is New"] = False
    df["First Seen"] = ""
    df["Last Seen"] = ""
    return df[EXPORT_MATCH_COLUMNS]


def matches_to_main_df(matches_df: pd.DataFrame) -> pd.DataFrame:
    if matches_df.empty:
        return pd.DataFrame(columns=EXPORT_MAIN_COLUMNS)

    key_cols = ["Event ID", "Timeslot Start (epoch)"]

    nearest = (
        matches_df.sort_values(by=["Distance to Property (miles)"], ascending=True)
        .groupby(key_cols, as_index=False)
        .first()
    )

    prop_lists = (
        matches_df.groupby(key_cols)
        .apply(
            lambda g: ", ".join(
                [
                    f"{r['Property ID']}:{r['Property']}({r['Distance to Property (miles)']}mi)"
                    for _, r in g.sort_values("Distance to Property (miles)").iterrows()
                ]
            )
        )
        .reset_index(name="Matched Properties (within radius)")
    )

    main_df = nearest.merge(prop_lists, on=key_cols, how="left").rename(
        columns={
            "Property": "Nearest Property",
            "Property ID": "Nearest Property ID",
            "Distance to Property (miles)": "Distance to Nearest Property (miles)",
        }
    )

    for col in ["Source", "Organization", "Event Key", "Is New", "First Seen"]:
        if col not in main_df.columns:
            main_df[col] = "" if col != "Is New" else False

    main_df = main_df[
        [
            "Source",
            "Organization",
            "Event Key",
            "Is New",
            "First Seen",
            "Protest Name",
            "Date",
            "Time",
            "Location",
            "Nearest Property",
            "Nearest Property ID",
            "Distance to Nearest Property (miles)",
            "Event URL",
            "Event Type",
            "Event ID",
            "Timeslot Start (epoch)",
            "Matched Properties (within radius)",
        ]
    ].sort_values(by=["Is New", "Distance to Nearest Property (miles)", "Date", "Time"], ascending=[False, True, True, True])

    return main_df


# -----------------------------
# Mobilize collection runner (zip-deduped)
# -----------------------------
def run_mobilize_collection(
    *,
    zip_groups: Dict[str, List[Property]],
    query_zips: List[str],
    now_epoch: int,
    end_epoch: int,
    api_max_dist: float,
    per_page: int,
    timeout: int,
    include_non_protest: bool,
    filter_cfg: Dict[str, Any],
    limiter: GlobalRateLimiter,
    workers: int,
    radius_miles: float,
    source_label: str,
    org_name: str,
    organization_id: Optional[int],
    show_progress: bool,
) -> pd.DataFrame:
    all_rows: List[Dict[str, Any]] = []
    total_zips = len(query_zips)
    done_zips = 0
    failed_zips = 0
    start_ts = time.time()

    with ThreadPoolExecutor(max_workers=max(1, workers)) as ex:
        futures = {
            ex.submit(
                query_mobilize_events_for_zip,
                z,
                now_epoch,
                end_epoch,
                api_max_dist,
                per_page,
                timeout,
                include_non_protest,
                filter_cfg,
                limiter,
                organization_id,
            ): z
            for z in query_zips
        }

        for fut in as_completed(futures):
            z = futures[fut]
            try:
                zip_events = fut.result()
                for ev in zip_events:
                    ep = (float(ev["event_lat"]), float(ev["event_lon"]))
                    for prop in zip_groups[z]:
                        pp = (prop.lat, prop.lon)
                        dist = compute_distance_miles(pp, ep)
                        if dist > radius_miles:
                            continue
                        all_rows.append(
                            {
                                "Source": source_label,
                                "Organization": org_name,
                                "Protest Name": ev["title"],
                                "Date": ev["date"],
                                "Time": ev["time"],
                                "Location": ev["location_str"],
                                "Property": prop.name,
                                "Property ID": prop.property_id,
                                "Distance to Property (miles)": round(dist, 2),
                                "Event URL": ev["browser_url"],
                                "Event Type": ev["event_type"],
                                "Event ID": ev["event_id"],
                                "Timeslot Start (epoch)": ev["timeslot_start"],
                            }
                        )
            except Exception as e:
                failed_zips += 1
                print(f"ZIP query failed: {z}: {e}", file=sys.stderr)
            finally:
                done_zips += 1
                if show_progress:
                    line = render_progress("ZIP queries", done_zips, total_zips, start_ts, failed_zips, z)
                    sys.stdout.write("\r" + line)
                    sys.stdout.flush()

    if show_progress:
        sys.stdout.write("\n")
        sys.stdout.flush()

    if not all_rows:
        df = pd.DataFrame(columns=EXPORT_MATCH_COLUMNS)
        return df

    df = pd.DataFrame(all_rows)
    # placeholders; seen will be applied later
    df["Event Key"] = ""
    df["Is New"] = False
    df["First Seen"] = ""
    df["Last Seen"] = ""
    return df[EXPORT_MATCH_COLUMNS]


def read_existing_excel(path: str, sheet_name: str, columns: List[str]) -> pd.DataFrame:
    try:
        df = pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl")
        for c in columns:
            if c not in df.columns:
                df[c] = pd.NA
        return df[columns]
    except FileNotFoundError:
        return pd.DataFrame(columns=columns)
    except ValueError:
        return pd.DataFrame(columns=columns)


# -----------------------------
# Main
# -----------------------------
def main() -> int:
    args = parse_args()

    props_df = read_properties(args.properties)

    if args.rewrite_properties:
        rewrite_properties(props_df, args.rewrite_properties)
        return 0

    if args.validate_properties is not None:
        bad_count = validate_properties_to_csv(props_df, args.validate_properties)
        if bad_count:
            print(f"Validation complete: {bad_count} bad row(s) written to {args.validate_properties}", file=sys.stderr)
            return 2
        print("Validation complete: 0 bad rows found.")
        return 0

    output_path = build_output_path(args)
    filter_cfg = load_filter_config(args.filter_config)
    limiter = GlobalRateLimiter(args.min_request_interval)

    props = normalize_properties(
        props_df,
        geocode_missing=args.geocode_missing,
        geocode_cache_path=args.geocode_cache,
        geocode_rate_seconds=args.geocode_rate_seconds,
    )
    zip_groups = group_properties_by_query_zip(props)
    query_zips = sorted(zip_groups.keys())

    now_ts = epoch_now()
    api_max_dist = args.radius_miles + max(0.0, args.api_buffer_miles)

    seen_store = load_seen_store(args.seen_store)

    # -------------------------
    # General (Mobilize) - 7 days default
    # -------------------------
    end_general = now_ts + args.window_days * 24 * 60 * 60
    print(f"[{dt.datetime.now().strftime('%H:%M:%S')}] Starting General Mobilize collection ({len(query_zips)} ZIPs)...", flush=True)
    general_matches_df = run_mobilize_collection(
        zip_groups=zip_groups,
        query_zips=query_zips,
        now_epoch=now_ts,
        end_epoch=end_general,
        api_max_dist=api_max_dist,
        per_page=args.per_page,
        timeout=args.timeout,
        include_non_protest=args.include_non_protest,
        filter_cfg=filter_cfg,
        limiter=limiter,
        workers=args.workers,
        radius_miles=args.radius_miles,
        source_label="General",
        org_name="",
        organization_id=None,
        show_progress=args.progress,
    )

    # Apply seen flags to matches + main
    print(f"[{dt.datetime.now().strftime('%H:%M:%S')}] General collection done. Applying seen flags...", flush=True)
    general_matches_df = apply_seen_flags(general_matches_df, seen_store, now_ts)
    general_main_df = matches_to_main_df(general_matches_df)

    # -------------------------
    # NoKings (Mobilize + Action Network) - 30 days default
    # -------------------------
    no_kings_main_df = pd.DataFrame(columns=EXPORT_MAIN_COLUMNS)
    if args.no_kings:
        end_nk = now_ts + args.no_kings_window_days * 24 * 60 * 60

        print(f"[{dt.datetime.now().strftime('%H:%M:%S')}] Starting NoKings Mobilize collection ({len(query_zips)} ZIPs, keyword filter)...", flush=True)
        nk_mobilize_matches = run_mobilize_collection(
            zip_groups=zip_groups,
            query_zips=query_zips,
            now_epoch=now_ts,
            end_epoch=end_nk,
            api_max_dist=api_max_dist,
            per_page=args.per_page,
            timeout=args.timeout,
            include_non_protest=True,
            filter_cfg=filter_cfg,
            limiter=limiter,
            workers=args.workers,
            radius_miles=args.radius_miles,
            source_label="NoKings-Mobilize",
            org_name="No Kings",
            organization_id=None,
            show_progress=args.progress,
        )
        if not nk_mobilize_matches.empty:
            mask = nk_mobilize_matches["Protest Name"].str.lower().str.contains("no kings", na=False)
            nk_mobilize_matches = nk_mobilize_matches[mask]
            print(f"[{dt.datetime.now().strftime('%H:%M:%S')}] NoKings keyword filter: kept {mask.sum()} of {len(mask)} events.", flush=True)

        print(f"[{dt.datetime.now().strftime('%H:%M:%S')}] NoKings Mobilize done. Starting Action Network scrape...", flush=True)
        seeds = [s.strip() for s in (args.action_network_seeds or "").split(",") if s.strip()] or DEFAULT_ACTION_NETWORK_SEEDS

        an_events = collect_action_network_events(
            seeds=seeds,
            now_epoch=now_ts,
            end_epoch=end_nk,
            timeout=args.timeout,
            max_seed_pages=args.action_network_max_seed_pages,
            max_event_links=args.action_network_max_event_links,
            show_progress=args.progress,
        )

        print(f"[{dt.datetime.now().strftime('%H:%M:%S')}] Action Network scrape done ({len(an_events)} events). Matching...", flush=True)
        nk_action_matches = build_matches_for_events(
            events=an_events,
            props=props,
            radius_miles=args.radius_miles,
            source_label="NoKings-ActionNetwork",
            org_name="No Kings (Action Network)",
            event_geocode_cache_path=args.event_geocode_cache,
            geocode_rate_seconds=args.geocode_rate_seconds,
            show_progress=args.progress,
        )

        # Apply seen flags to matches before combining
        nk_mobilize_matches = apply_seen_flags(nk_mobilize_matches, seen_store, now_ts)
        nk_action_matches = apply_seen_flags(nk_action_matches, seen_store, now_ts)

        no_kings_matches_all = pd.concat([nk_mobilize_matches, nk_action_matches], ignore_index=True)
        if not no_kings_matches_all.empty:
            no_kings_matches_all = no_kings_matches_all.drop_duplicates(
                subset=["Event URL", "Timeslot Start (epoch)", "Property ID"], keep="first"
            )

        no_kings_main_df = matches_to_main_df(no_kings_matches_all)

        if not no_kings_main_df.empty:
            no_kings_main_df = no_kings_main_df.drop_duplicates(
                subset=["Event URL", "Timeslot Start (epoch)"], keep="first"
            )

    # Persist seen store (first_seen ever, last_seen this run)
    save_seen_store(args.seen_store, seen_store)

    # -------------------------
    # Write workbook
    # -------------------------
    def _prep_for_output(df: pd.DataFrame, hide: set, dist_col: str) -> pd.DataFrame:
        out = df.drop(columns=[c for c in hide if c in df.columns])
        if dist_col in out.columns:
            out = out.sort_values(dist_col, ascending=True, na_position="last").reset_index(drop=True)
        return out

    print("Writing Excel workbook...", flush=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        main_out = mark_duplicates(_prep_for_output(general_main_df, _OUTPUT_HIDE_MAIN, "Distance to Nearest Property (miles)"))
        main_out.to_excel(writer, index=False, sheet_name=DEFAULT_SHEET_MAIN)
        if not args.no_autosize:
            autosize_worksheet_columns(writer, DEFAULT_SHEET_MAIN)
        apply_sheet_formatting(writer, DEFAULT_SHEET_MAIN)
        if not args.no_highlight_new:
            apply_highlight_new_rows(writer, DEFAULT_SHEET_MAIN)

        if args.keep_all_matches:
            matches_out = mark_duplicates(_prep_for_output(general_matches_df, _OUTPUT_HIDE_MATCH, "Distance to Property (miles)"))
            matches_out.to_excel(writer, index=False, sheet_name=DEFAULT_SHEET_MATCHES)
            if not args.no_autosize:
                autosize_worksheet_columns(writer, DEFAULT_SHEET_MATCHES)
            apply_sheet_formatting(writer, DEFAULT_SHEET_MATCHES)
            if not args.no_highlight_new:
                apply_highlight_new_rows(writer, DEFAULT_SHEET_MATCHES)
            # Hide the sheet so it exists for reference but isn't visible by default
            writer.sheets[DEFAULT_SHEET_MATCHES].sheet_state = "hidden"

        if args.no_kings:
            nk_out = mark_duplicates(_prep_for_output(no_kings_main_df, _OUTPUT_HIDE_MAIN, "Distance to Nearest Property (miles)"))
            nk_out.to_excel(writer, index=False, sheet_name=DEFAULT_SHEET_NO_KINGS)
            if not args.no_autosize:
                autosize_worksheet_columns(writer, DEFAULT_SHEET_NO_KINGS)
            apply_sheet_formatting(writer, DEFAULT_SHEET_NO_KINGS)
            if not args.no_highlight_new:
                apply_highlight_new_rows(writer, DEFAULT_SHEET_NO_KINGS)

    print(f"Wrote General rows: {len(general_main_df)} -> {output_path}")
    if args.no_kings:
        print(f"Wrote NoKings rows: {len(no_kings_main_df)} -> sheet '{DEFAULT_SHEET_NO_KINGS}'")
    print(f"Seen store updated: {args.seen_store}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
